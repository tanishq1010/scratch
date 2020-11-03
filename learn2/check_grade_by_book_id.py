import os
import csv
import json
import string
import random
import requests
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook
import ast


class Source(object):
    def __init__(self):
        super(Source, self).__init__()
        self.headers = {
            'Connection': 'keep-alive',
            'Accept': '*/*',
            'Authorization': '048f38be-1b07-4b21-8f24-eac727dce217:gSEkC3dqDcIv1bbOk78UD9owjn7ins8D',
            'Content-Type': 'application/json',
        }

    def callAPI2(self, url, payload, method):
        response = requests.request(method, url, headers=self.headers, data=payload)
        if response.status_code != 200:
            print(url + ' - ' + str(response.content))
            return None
        return response

    def main(self):

        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Exam_Name"
        sheet["B1"] = "Book_id"
        sheet["C1"] = "Book_Name"
        sheet["D1"] = "Authors"

        workbook.save(filename="CG_DB_Book_data.xlsx")
        response1 = self.callAPI2(
            'https://content-demo.embibe.com/learning_map_formats?where={"status":"active","type":"book"}',
            "{}", 'GET')

        try:
            print(response1.status_code)
        except:
            print("learning_map_formats?where= API gave error response")
            os.remove("CG_DB_Book_data.xlsx")
            outF = open("txtfile.txt", "w")
            outF.write("learning_map_formats?where= API gave error response")
            sys.exit()

        response1 = self.callAPI2(
            'https://content-demo.embibe.com/learning_map_formats?where={"status":"active","type":"book"}&max_results=' + str(
                response1.json()["_meta"]["total"] + 1), "{}", 'GET')
        i = int(0)

        if response1.status_code == 200:
            print("\tStatus: ", response1.status_code)
            print("\t\tSaving in CG_DB_Book_data.xlsx")
            for goal in response1.json()["_items"]:
                home_data = []
                wb = load_workbook("CG_DB_Book_data.xlsx")
                sheet = wb["Sheet"]
                # home_data.append(str(goal["content"]["grade"]))
                length = goal["content"]["grade"]
                # print(goal["content"]["grade"][0])
                for j in range(0, len(length)):
                    # length2 = goal["content"]["authors"]
                    # for k in range(0, len(length2)):
                    try:

                        home_data = []
                        home_data.append(goal["content"]["grade"][j])
                        home_data.append(str(goal["_id"]))
                        s = str(goal["display_name"])
                        home_data.append(s.upper())
                        try:
                            z = str(goal["content"]["authors"])
                        except Exception as e:
                            print(e)
                            z = ""
                        home_data.append(z.upper())
                        sheet.append(home_data)

                    except Exception as e:
                        print(e)

                wb.save(filename="CG_DB_Book_data.xlsx")

                print(i)
                i += 1
                # if i > 9:
                #     break
        else:
            print("learning_map_formats?where= API response status code != 200 ")
            print(response1.status_code)

    def get_books_of_exam(self, exam, df):
        return df[df['Exam_Name'].str.contains(exam)]


def check_grade_by_book_id(book_id, exam):
    src = Source()

    if not os.path.exists("CG_DB_Book_data.xlsx"):
        print("\tCG_DB_Book_data.xlsx not found. creating new one")
        src.main()
    else:
        print("\tCG_DB_Book_data.xlsx found. Reading.....")
    #         src.main()

    df = pd.read_excel("CG_DB_Book_data.xlsx")
    df = df.loc[df["Book_id"] == book_id]

    #     print(df)

    if len(df) == 0:
        return "NA"
    else:
        df = src.get_books_of_exam(exam, df)
        if len(df) > 0:
            return "Yes"
        else:
            return "No"